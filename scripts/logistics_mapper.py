import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.chart.label import DataLabelList
import logging
import sys

class HVDCLogisticsMapper:
    def __init__(self):
        self.data_dir = Path('data')
        self.output_dir = Path('output')
        self.log_dir = Path('logs')
        
        # 디렉토리 생성
        for dir_path in [self.data_dir, self.output_dir, self.log_dir]:
            dir_path.mkdir(parents=True, exist_ok=True)
        
        # 로깅 설정
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(self.log_dir / 'logistics_mapper.log', encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
    
    def determine_step(self, row):
        """현재 프로세스 단계 결정 (1~5)"""
        if pd.notna(row["MOSB"]):
            return 5  # 현장 수령 완료
        if pd.notna(row["DSV\n Outdoor"]):
            return 4  # 운송 중
        if pd.notna(row["Customs\n Start"]):
            return 3  # 통관 완료, 창고 입고 단계
        if pd.notna(row["ATA"]):
            return 2  # 입항 후 통관 진행
        return 1  # 해외 조달중 (UAE 도착 전)
    
    def detect_site(self, row):
        """현장(SITE) 감지 및 FLOW 분류 - MIR·SHU=육상, AGI·DAS=섬, 미식별시 UNK"""
        site_cols = ["MIR", "SHU", "DAS", "AGI"]
        for col in site_cols:
            if col in row and pd.notna(row[col]):
                if col in ["MIR", "SHU"]:
                    return col  # 육상
                elif col in ["AGI", "DAS"]:
                    return col  # 섬
        self.logger.warning(f"SITE 미식별: {row.get('NO.', '')}")
        return "UNK"
    
    def refined_hvdc_step(self, desc: str) -> int:
        """정교화된 HVDC 공정단계 분류 함수 (v3)"""
        if not isinstance(desc, str) or not desc.strip():
            return 99
        d = desc.lower()

        # 1. Converter
        if any(k in d for k in [
            "converter", "valve hall", "thyristor", "igbt", "converter transformer",
            "transformer", "synchronous condenser", "valve support", "cooling unit",
            "valve", "power electronics", "semiconductor"
        ]):
            return 1

        # 2. Transmission
        if any(k in d for k in [
            "dc cable", "submarine cable", "transmission line", "power cable", "conductor",
            "cable termination", "sealing end", "joint", "gis termination", "conduit",
            "busbar", "aluminium busbar", "copper busbar", "bus bar", "bus-bar",
            "cable", "wire", "conductor", "transmission", "power line"
        ]):
            return 2

        # 3. Filter / Reactor
        if any(k in d for k in [
            "ac harmonic filter", "dc filter", "harmonic filter", "smoothing reactor",
            "high frequency filter", "reactor", "capacitor", "lc filter", "ac filter", "insulator", "gis",
            "layer", "layers", "filter layer", "filter bank", "filter unit",
            "reactor bank", "reactor unit", "capacitor bank", "capacitor unit"
        ]):
            return 3

        # 4. Control / Protection
        if any(k in d for k in [
            "scada", "control system", "relay", "protection panel", "monitoring",
            "plc", "hmi", "rtu", "breaker", "busbar protection", "system",
            "control cubicle", "control panel", "control unit", "control box",
            "optical", "fiber", "fibre", "transducer", "sensor", "transmitter",
            "battery", "batteries", "power supply", "ups", "dc power",
            "cubicle", "panel", "cabinet", "enclosure", "housing"
        ]):
            return 4

        # 5. Grounding / Earthing
        if any(k in d for k in [
            "grounding", "earth electrode", "electrode line", "neutral bus", "nbgs", "nbs",
            "earthing", "ground", "earth", "neutral", "grounding system",
            "grounding rod", "grounding wire", "grounding cable"
        ]):
            return 5

        # 6. Spare / Maintenance
        if any(k in d for k in [
            "spare", "repair", "tool", "maintenance", "accessories", "consumable", "dummy",
            "spare part", "replacement", "backup", "reserve", "standby",
            "maintenance kit", "repair kit", "tool kit", "tool set"
        ]):
            return 6

        return 99
    
    def add_container_summary(self, df, writer):
        """컨테이너 데이터 집계"""
        ship_col = "SCT SHIP NO."
        if ship_col not in df.columns:
            self.logger.warning("'SCT SHIP NO.' 컬럼이 없어 Container_Summary 시트 생략")
            return

        # 컨테이너 관련 컬럼 정의
        container_cols = {
            '20ft': ['20DC', '20OT(IN)', '20OT(OH)', '20FR(IN)', '20FR(FV)', '20FR(OW)', '20FR(OW,OH)'],
            '40ft': ['40DC', '40HQ', '45HQ', '40OT(IN)', '40OT(OH)', '40FR(IN)', '40FR(OW)', '40FR(OW,OH)', '40FR(OW,OL)']
        }
        
        # 숫자형 변환 및 합계 계산
        for col in ['PKG'] + container_cols['20ft'] + container_cols['40ft'] + ['QTY OF CNTR']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            else:
                df[col] = 0
        
        # 20ft, 40ft 컨테이너 합계 계산
        df['20FT_SUM'] = df[container_cols['20ft']].sum(axis=1)
        df['40FT_SUM'] = df[container_cols['40ft']].sum(axis=1)
        
        # SCT SHIP NO.별 집계
        ship_summary = df.groupby(ship_col).agg({
            'PKG': 'sum',
            '20FT_SUM': 'sum',
            '40FT_SUM': 'sum',
            'QTY OF CNTR': 'sum'
        }).rename(columns={
            'PKG': '총 패키지 수',
            '20FT_SUM': '20ft 컨테이너 수',
            '40FT_SUM': '40ft 컨테이너 수',
            'QTY OF CNTR': '전체 컨테이너 수'
        })
        
        # 결과를 엑셀에 저장
        ship_summary.to_excel(writer, sheet_name="Container_Summary")
        
        # 시트 포맷팅
        workbook = writer.book
        worksheet = workbook["Container_Summary"]
        
        # 헤더 스타일 설정
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 데이터 정렬
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
        
        # 컬럼 너비 조정
        worksheet.column_dimensions['A'].width = 20  # SCT SHIP NO.
        worksheet.column_dimensions['B'].width = 15  # 총 패키지 수
        worksheet.column_dimensions['C'].width = 15  # 20ft 컨테이너 수
        worksheet.column_dimensions['D'].width = 15  # 40ft 컨테이너 수
        worksheet.column_dimensions['E'].width = 15  # 전체 컨테이너 수
    
    def create_excel_report(self, df, output_file):
        """종합 엑셀 리포트 생성"""
        self.logger.info("엑셀 리포트 생성 시작")
        
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 1. STEP_FLOW 시트
                df.to_excel(writer, sheet_name='STEP_FLOW', index=False)
                
                # 2. SLA_Exceed 시트
                sla_exceeded = df[
                    (df["입항→통관"] > 3) |
                    (df["통관→창고출고"] > 2) |
                    (df["창고출고→현장도착"] > 5) |
                    (df["전체 리드타임"] > 30)
                ].copy()
                sla_exceeded.to_excel(writer, sheet_name='SLA_Exceed', index=False)
                
                # 3. Process_Vendor 시트
                process_vendor = df.groupby(['STEP_NAME', 'VENDOR'])['전체 리드타임'].agg([
                    'count', 'mean', 'min', 'max'
                ]).round(1)
                process_vendor.to_excel(writer, sheet_name='Process_Vendor')
                
                # 4. Site_Summary 시트
                site_summary = df.groupby('SITE')['전체 리드타임'].agg([
                    'count', 'mean', 'min', 'max'
                ]).round(1)
                site_summary.to_excel(writer, sheet_name='Site_Summary')
                
                # 5. MOSB_Pred 시트
                mosb_pred = df[df['STEP_NO'] < 5].copy()
                mosb_pred['예상_MOSB'] = mosb_pred.apply(
                    lambda x: x['DSV\n Outdoor'] + timedelta(days=5) if pd.notna(x['DSV\n Outdoor'])
                    else x['Customs\n Start'] + timedelta(days=7) if pd.notna(x['Customs\n Start'])
                    else x['ATA'] + timedelta(days=10) if pd.notna(x['ATA'])
                    else None,
                    axis=1
                )
                mosb_pred = mosb_pred[['NO.', 'VENDOR', 'SUB DESCRIPTION', 'STEP_NAME', 
                                    'SITE', 'FLOW', 'DSV\n Outdoor', '예상_MOSB']]
                mosb_pred.to_excel(writer, sheet_name='MOSB_Pred', index=False)
                
                # 6. Dashboard 시트
                dashboard = pd.DataFrame({
                    '지표': [
                        '총 자재 수',
                        '현장 도착 완료',
                        '운송 중',
                        '통관 완료',
                        '입항 완료',
                        '해외 조달중',
                        '평균 리드타임',
                        'SLA 초과 건수'
                    ],
                    '값': [
                        len(df),
                        len(df[df['STEP_NO'] == 5]),
                        len(df[df['STEP_NO'] == 4]),
                        len(df[df['STEP_NO'] == 3]),
                        len(df[df['STEP_NO'] == 2]),
                        len(df[df['STEP_NO'] == 1]),
                        df['전체 리드타임'].mean(),
                        len(sla_exceeded)
                    ]
                })
                dashboard.to_excel(writer, sheet_name='Dashboard', index=False)
                
                # 7. Container_Summary 시트
                self.add_container_summary(df, writer)
                
                # 8. 통계 정보 저장
                stats_file = self.output_dir / 'logistics_stats.txt'
                with open(stats_file, 'w', encoding='utf-8') as f:
                    f.write("=== HVDC 물류 통계 ===\n\n")
                    f.write(f"총 자재 수: {len(df)}\n")
                    f.write(f"현장 도착 완료: {len(df[df['STEP_NO'] == 5])}\n")
                    f.write(f"운송 중: {len(df[df['STEP_NO'] == 4])}\n")
                    f.write(f"통관 완료: {len(df[df['STEP_NO'] == 3])}\n")
                    f.write(f"입항 완료: {len(df[df['STEP_NO'] == 2])}\n")
                    f.write(f"해외 조달중: {len(df[df['STEP_NO'] == 1])}\n\n")
                    f.write(f"평균 리드타임: {df['전체 리드타임'].mean():.1f}일\n")
                    f.write(f"최대 리드타임: {df['전체 리드타임'].max():.1f}일\n")
                    f.write(f"최소 리드타임: {df['전체 리드타임'].min():.1f}일\n\n")
                    f.write(f"SLA 초과 건수: {len(sla_exceeded)}\n")
                
                self.logger.info(f"통계 정보 저장 완료: {stats_file}")
            
            self.logger.info(f"엑셀 리포트 생성 완료: {output_file}")
            
        except Exception as e:
            self.logger.error(f"엑셀 리포트 생성 중 오류 발생: {str(e)}")
            raise
    
    def process_data(self, df):
        """데이터 전처리 및 매핑"""
        self.logger.info("데이터 전처리 및 매핑 시작")
        
        try:
            # 1. 날짜 컬럼 변환
            date_cols = ['ATA', 'Customs\n Start', 'DSV\n Outdoor', 'MOSB', 'AAA Storage', 'ZENER (WH)', 'Hauler DG Storage', 'Vijay Tanks']
            for col in date_cols:
                if col in df.columns:
                    df[col] = pd.to_datetime(df[col], errors='coerce')
            
            # 2. 프로세스 단계 결정
            df['STEP_NO'] = df.apply(self.determine_step, axis=1)
            df['STEP_NAME'] = df['STEP_NO'].map({
                1: '해외 조달중',
                2: '입항 완료',
                3: '통관 완료',
                4: '운송 중',
                5: '현장 도착 완료'
            })
            
            # 3. 현장 감지
            df['SITE'] = df.apply(self.detect_site, axis=1)
            df['FLOW'] = df['SITE'].map({
                'MIR': 'MIR',
                'SHU': 'SHU',
                'DAS': 'DAS',
                'AGI': 'AGI'
            })
            
            # 4. 공정단계 분류
            df['공정단계_HVDC'] = df['SUB DESCRIPTION'].apply(self.refined_hvdc_step)
            df['공정단계_HVDC_Label'] = df['공정단계_HVDC'].map({
                1: 'Converter',
                2: 'Transmission',
                3: 'Filter/Reactor',
                4: 'Control/Protection',
                5: 'Grounding',
                6: 'Spare/Maintenance',
                99: '기타'
            })
            
            # 5. 리드타임 계산
            df['입항→통관'] = (df['Customs\n Start'] - df['ATA']).dt.days
            df['통관→창고출고'] = (df['DSV\n Outdoor'] - df['Customs\n Start']).dt.days
            df['창고출고→현장도착'] = (df['MOSB'] - df['DSV\n Outdoor']).dt.days
            df['전체 리드타임'] = (df['MOSB'] - df['ATA']).dt.days
            # 섬 운송(SITE=AGI/DAS) +5일 보정
            island_mask = df['SITE'].isin(['AGI', 'DAS'])
            df.loc[island_mask, '입항→통관'] = df.loc[island_mask, '입항→통관'] + 5
            
            # 6. 컨테이너 데이터 처리
            if '20ft Q\'TY' in df.columns and '40ft Q\'TY' in df.columns:
                df['20ft Q\'TY'] = pd.to_numeric(df['20ft Q\'TY'], errors='coerce').fillna(0)
                df['40ft Q\'TY'] = pd.to_numeric(df['40ft Q\'TY'], errors='coerce').fillna(0)
                df['TOTAL Q\'TY'] = df['20ft Q\'TY'] + df['40ft Q\'TY']
            
            self.logger.info("데이터 전처리 및 매핑 완료")
            return df
            
        except Exception as e:
            self.logger.error(f"데이터 처리 중 오류 발생: {str(e)}")
            raise
    
    def run_pipeline(self):
        """전체 파이프라인 실행"""
        try:
            # 1. 데이터 로드
            data_path = self.data_dir / 'HVDC-STATUS.xlsx'
            if not data_path.exists():
                raise FileNotFoundError(f"원본 데이터 파일을 찾을 수 없습니다: {data_path}")
            
            df = pd.read_excel(data_path)
            self.logger.info(f"원본 데이터 로드 완료: {len(df)} 행")
            
            # 2. 데이터 처리
            df = self.process_data(df)
            
            # 3. 결과 저장
            output_file = self.output_dir / 'logistics_mapping.xlsx'
            self.create_excel_report(df, output_file)
            
            self.logger.info("=== 전체 파이프라인 완료 ===")
            return True
            
        except Exception as e:
            self.logger.error(f"파이프라인 실행 중 오류 발생: {str(e)}")
            return False

def main():
    mapper = HVDCLogisticsMapper()
    if not mapper.run_pipeline():
        sys.exit(1)

if __name__ == "__main__":
    main() 